import os
import uuid
import hashlib
from datetime import datetime, UTC, date, time
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


def serialize_cell_value(cell_value):
    """Convert cell values to JSON-serializable format."""
    if cell_value is None:
        return ""
    elif isinstance(cell_value, (datetime, date)):
        return cell_value.isoformat()
    elif isinstance(cell_value, time):
        return cell_value.isoformat()
    else:
        return str(cell_value)


def convert_excel_to_gsheets_format(file_path, file_name=None):
    """Convert an Excel file to Google Sheets format JSON."""
    if file_name is None:
        file_name = os.path.basename(file_path)
    
    file_extension = os.path.splitext(file_name)[1].lower()
    file_stats = os.stat(file_path)
    spreadsheet_id = f"sheet_{hashlib.md5(file_path.encode()).hexdigest()}"
    current_time = datetime.fromtimestamp(file_stats.st_mtime, UTC).strftime('%Y-%m-%dT%H:%M:%SZ')
    
    # Create base metadata following SheetsDefaultDB.json format exactly
    # This matches the Drive file structure but with Sheets-specific fields
    json_data = {
        "id": spreadsheet_id,
        "driveId": "",
        "name": file_name,
        "mimeType": "application/vnd.google-apps.spreadsheet",
        "createdTime": datetime.fromtimestamp(file_stats.st_ctime, UTC).strftime('%Y-%m-%dT%H:%M:%SZ'),
        "modifiedTime": current_time,
        "parents": [],
        "owners": ["john.doe@gmail.com"],
        "size": 0,  # Sheets typically show 0 size
        "trashed": False,
        "starred": False,
        "properties": {
            "title": os.path.splitext(file_name)[0],
            "locale": "en_US",
            "timeZone": "UTC"
        },
        "sheets": [],
        "data": {},
        "permissions": [
            {
                "id": f"permission_{spreadsheet_id}",
                "role": "owner",
                "type": "user",
                "emailAddress": "john.doe@gmail.com"
            }
        ]
    }
    
    try:
        wb = load_workbook(file_path, read_only=True, data_only=True)
        
        for index, sheet_name in enumerate(wb.sheetnames):
            ws = wb[sheet_name]
            sheet_id = index + 1
            
            # Find actual data range
            min_row, min_col, max_row, max_col = None, None, 0, 0
            
            for r_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
                row_has_data = False
                for c_idx, cell in enumerate(row, start=1):
                    if cell is not None and str(cell).strip() != "":
                        row_has_data = True
                        if min_row is None or r_idx < min_row:
                            min_row = r_idx
                        if min_col is None or c_idx < min_col:
                            min_col = c_idx
                        if c_idx > max_col:
                            max_col = c_idx
                if row_has_data and r_idx > max_row:
                    max_row = r_idx
            
            # Add sheet metadata matching the expected format
            json_data["sheets"].append({
                "properties": {
                    "sheetId": sheet_id,
                    "title": sheet_name,
                    "index": index,
                    "sheetType": "GRID",
                    "gridProperties": {
                        "rowCount": max(ws.max_row or 1000, 1000),  # Use at least 1000 rows
                        "columnCount": max(ws.max_column or 26, 26)  # Use at least 26 columns
                    }
                }
            })
            
            # Extract data using A1 notation format: SheetName!A1:C10
            if all([min_row, min_col, max_row, max_col]):
                # Convert to A1 notation format
                start_col_letter = get_column_letter(min_col)
                end_col_letter = get_column_letter(max_col)
                range_key = f"{sheet_name}!{start_col_letter}{min_row}:{end_col_letter}{max_row}"
                sheet_data = []
                for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col, values_only=True):
                    sheet_data.append([serialize_cell_value(cell) for cell in row])
                json_data["data"][range_key] = sheet_data
        
        wb.close()
        
    except Exception as e:
        json_data["error"] = f"Error reading Excel file: {e}"
    
    return json_data 